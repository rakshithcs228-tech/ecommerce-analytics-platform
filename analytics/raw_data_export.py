"""
╔══════════════════════════════════════════════════════════════════╗
║  E-COMMERCE ANALYTICS PLATFORM — RAW DATA EXPORT                ║
║                                                                  ║
║  Exports full raw data from all Delta Lake tables to Excel.      ║
║                                                                  ║
║  Sheets:                                                         ║
║    1.  Index                — Table overview & row counts        ║
║    2.  Orders_Raw           — Full curated orders table          ║
║    3.  Payments_Raw         — Full curated payments table        ║
║    4.  User_Events_Raw      — Full curated user_events table     ║
║    5.  Customer_RFM_Raw     — Full customer_rfm summary          ║
║    6.  Churn_Candidates_Raw — Full churn_candidates summary      ║
║                                                                  ║
║  Output:                                                         ║
║    → Ecommerce_Raw_Data.xlsx (same folder as this script)        ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ── STEP 1: ENVIRONMENT SETUP ─────────────────────────────────────
import os
import re
import sys
import json
from datetime import datetime

os.environ["HADOOP_HOME"]       = r"C:\hadoop"
os.environ["PATH"]              = r"C:\hadoop\bin;" + os.environ.get("PATH", "")
os.environ["USERPROFILE"]       = r"D:\spark-cache"
os.environ["HOME"]              = r"D:\spark-cache"
os.environ["IVY_HOME"]          = r"D:\spark-cache\.ivy2"
os.environ["JAVA_TOOL_OPTIONS"] = "-Djava.io.tmpdir=D:\\spark-tmp"

# ── STEP 2: IMPORTS ───────────────────────────────────────────────
from pyspark.sql import SparkSession
from pyspark.sql import functions as F
from pyspark.sql.types import (
    TimestampType, DateType, ArrayType, StructType, MapType
)

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── STEP 3: PATHS ─────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
BASE_PATH    = os.path.join(PROJECT_ROOT, "delta-lake")

CURATED_PATH = os.path.join(BASE_PATH, "curated")
SUMMARY_PATH = os.path.join(BASE_PATH, "summaries")
EXCEL_PATH   = os.path.join(SCRIPT_DIR, "Ecommerce_Raw_Data.xlsx")

# ── STEP 4: STYLE CONSTANTS ───────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL     = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
BODY_FONT    = Font(name="Arial", size=9)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    top=Side(style="thin",    color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Excel limits
MAX_EXCEL_ROWS = 1048576
MAX_CELL_LENGTH = 32767

# Remove characters that Excel XML cannot store
ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0B\x0C\x0E-\x1F]'
)


# ══════════════════════════════════════════════════════════════════
# SPARK SESSION
# ══════════════════════════════════════════════════════════════════
def create_spark():
    spark = (
        SparkSession.builder
        .appName("EcommerceRawDataExport")
        .master("local[*]")
        .config("spark.jars.packages",
                "io.delta:delta-core_2.12:2.4.0")
        .config("spark.sql.extensions",
                "io.delta.sql.DeltaSparkSessionExtension")
        .config("spark.sql.catalog.spark_catalog",
                "org.apache.spark.sql.delta.catalog.DeltaCatalog")
        .config("spark.sql.shuffle.partitions", "4")
        .config("spark.driver.memory", "4g")
        .config("spark.local.dir", "D:\\spark-tmp")
        .config("spark.ui.port", "4042")
        .config("spark.sql.execution.arrow.pyspark.enabled", "false")
        .getOrCreate()
    )
    spark.sparkContext.setLogLevel("ERROR")
    return spark


# ══════════════════════════════════════════════════════════════════
# TABLE REGISTRY
# ══════════════════════════════════════════════════════════════════
TABLE_REGISTRY = [
    (os.path.join(CURATED_PATH, "orders"),
     "orders",           "Orders",           "curated"),
    (os.path.join(CURATED_PATH, "payments"),
     "payments",         "Payments",         "curated"),
    (os.path.join(CURATED_PATH, "user_events"),
     "user_events",      "User Events",      "curated"),
    (os.path.join(SUMMARY_PATH, "customer_rfm"),
     "customer_rfm",     "Customer RFM",     "summaries"),
    (os.path.join(SUMMARY_PATH, "churn_candidates"),
     "churn_candidates", "Churn Candidates", "summaries"),
]


def load_tables(spark):
    """Register Delta tables as SQL temp views. Returns {view: row_count}."""
    print("\n📂 Loading Delta Lake tables...")
    counts = {}
    for path, view, label, layer in TABLE_REGISTRY:
        delta_log = os.path.join(path, "_delta_log")
        if not os.path.exists(path) or not os.path.exists(delta_log):
            print(f"   ⚠️  [{layer}] {label} — not found, skipping")
            counts[view] = 0
            continue
        try:
            spark.read.format("delta").load(path) \
                 .createOrReplaceTempView(view)
            n = spark.sql(
                f"SELECT COUNT(*) AS n FROM {view}"
            ).collect()[0]["n"]
            print(f"   ✅ [{layer}] {label}: {n:,} rows")
            counts[view] = n
        except Exception as e:
            print(f"   ❌ [{layer}] {label} failed: {e}")
            counts[view] = 0
    return counts


# ══════════════════════════════════════════════════════════════════
# SAFE toPandas
# ══════════════════════════════════════════════════════════════════
def _is_complex(dtype):
    """Return True if a Spark DataType is Array, Struct, or Map."""
    return isinstance(dtype, (ArrayType, StructType, MapType))


def safe_to_pandas(sdf):
    """
    Convert a Spark DataFrame to a pandas DataFrame that openpyxl
    can serialise without errors.

    Handles three problem column types — all converted in Spark
    before toPandas() is called:

      1. TimestampType  → STRING  ("2026-05-27 08:39:40")
         then re-parsed as tz-naive datetime in pandas

      2. DateType       → STRING  ("2026-05-27")
         then re-parsed as Python date in pandas

      3. ArrayType / StructType / MapType  → JSON STRING
         e.g. items array of structs becomes a readable JSON string
         that fits neatly in an Excel cell
    """
    ts_cols      = []
    date_cols    = []
    complex_cols = []

    for field in sdf.schema.fields:
        if isinstance(field.dataType, TimestampType):
            ts_cols.append(field.name)
        elif isinstance(field.dataType, DateType):
            date_cols.append(field.name)
        elif _is_complex(field.dataType):
            complex_cols.append(field.name)

    # ── Cast in Spark ──────────────────────────────────────────────
    for c in ts_cols + date_cols:
        sdf = sdf.withColumn(c, F.col(c).cast("string"))

    for c in complex_cols:
        # to_json turns any Array/Struct/Map column into a JSON string
        sdf = sdf.withColumn(c, F.to_json(F.col(c)))

    if complex_cols:
        print(f"      ℹ️  Nested columns serialised to JSON: "
              f"{', '.join(complex_cols)}")

    # ── toPandas ──────────────────────────────────────────────────
    pdf = sdf.toPandas()

    # ── Re-parse date/timestamp strings ───────────────────────────
    for c in ts_cols:
        pdf[c] = pd.to_datetime(pdf[c], errors="coerce")

    for c in date_cols:
        pdf[c] = pd.to_datetime(pdf[c], errors="coerce").dt.date

    # ── Clean up residual None/nan strings ────────────────────────
    pdf = pdf.replace({"None": None, "nan": None})

    return pdf


# ══════════════════════════════════════════════════════════════════
# OPENPYXL HELPERS
# ══════════════════════════════════════════════════════════════════
def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER


def auto_col_widths(ws, df):
    for i, col_name in enumerate(df.columns, start=1):
        sample = df[col_name].astype(str).head(200)
        max_w  = sample.str.len().max() if len(sample) else 0
        width  = min(max(len(str(col_name)), int(max_w or 0)) + 3, 45)
        ws.column_dimensions[get_column_letter(i)].width = width


def sanitise_value(value):
    """
    Convert a cell value to something Excel/openpyxl can safely write.
    """

    if value is None:
        return None

    if isinstance(value, (np.integer,)):
        return int(value)

    if isinstance(value, (np.floating,)):
        return None if np.isnan(value) else float(value)

    if isinstance(value, (np.bool_,)):
        return bool(value)

    if isinstance(value, float) and np.isnan(value):
        return None

    if value is pd.NaT:
        return None

    # Convert complex structures to JSON
    if isinstance(value, (list, dict)):
        try:
            value = json.dumps(value, default=str)
        except Exception:
            value = str(value)

    # Handle Spark Row objects
    elif hasattr(value, "__class__") and "Row" in type(value).__name__:
        try:
            value = json.dumps(dict(value.asDict()), default=str)
        except Exception:
            value = str(value)

    # Clean strings for Excel
    if isinstance(value, str):
        # Remove illegal XML characters
        value = ILLEGAL_XML_RE.sub("", value)

        # Excel max cell length
        if len(value) > MAX_CELL_LENGTH:
            value = value[:MAX_CELL_LENGTH]

    return value
def write_dataframe(ws, df, start_row=1):
    """Write DataFrame to worksheet with header styling + alt-row fill."""
    # Header
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=col_idx, value=col_name)
    style_header_row(ws, start_row, len(df.columns))

    # Rows
    for row_offset, row_vals in enumerate(df.itertuples(index=False), start=1):
        excel_row = start_row + row_offset
        fill = ALT_FILL if row_offset % 2 == 0 else None
        for col_idx, value in enumerate(row_vals, start=1):
            value = sanitise_value(value)
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if fill:
                cell.fill = fill

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(len(df.columns))}{start_row}"
    )
    auto_col_widths(ws, df)


# ══════════════════════════════════════════════════════════════════
# INDEX SHEET
# ══════════════════════════════════════════════════════════════════
def write_index_sheet(wb, table_meta, export_time):
    ws = wb.active
    ws.title = "Index"

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "E-COMMERCE ANALYTICS PLATFORM — RAW DATA EXPORT"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    c.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = (f"Generated: {export_time}   |   "
                   "Full raw records from Delta Lake — no aggregations")
    c.font      = Font(name="Arial", size=9, italic=True, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    c.alignment = CENTER_ALIGN

    # Column headers
    headers = ["Sheet Name", "Delta Layer", "Table",
               "Description", "Row Count", "Columns"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER

    DESCRIPTIONS = {
        "orders":
            "All order transactions — amounts, items (JSON), device, region, value segment, flags",
        "payments":
            "All payment records — method, gateway, status, success flag, amounts",
        "user_events":
            "Clickstream / behavioural events — page views, add-to-cart, checkout, purchase",
        "customer_rfm":
            "Computed RFM scores and segments per customer",
        "churn_candidates":
            "Customers flagged as at-risk of churning with supporting metrics",
    }
    LAYERS = {
        "orders": "curated", "payments": "curated",
        "user_events": "curated",
        "customer_rfm": "summaries", "churn_candidates": "summaries",
    }
    SHEET_NAMES = {
        "orders": "Orders_Raw", "payments": "Payments_Raw",
        "user_events": "User_Events_Raw",
        "customer_rfm": "Customer_RFM_Raw",
        "churn_candidates": "Churn_Candidates_Raw",
    }

    for row_offset, (view, meta) in enumerate(table_meta.items(), start=1):
        excel_row = 4 + row_offset
        fill = ALT_FILL if row_offset % 2 == 0 else None
        row_data = [
            SHEET_NAMES.get(view, view),
            LAYERS.get(view, "—"),
            view,
            DESCRIPTIONS.get(view, "—"),
            meta["row_count"] if meta["available"] else "N/A",
            meta["col_count"] if meta["available"] else "—",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = BODY_FONT; cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if fill:
                cell.fill = fill
            if col_idx == 5 and meta["available"]:
                cell.font = Font(name="Arial", size=9,
                                 bold=True, color="1F6E43")

    col_widths = [22, 14, 22, 68, 14, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 18
    print("   ✅ Index sheet written")


# ══════════════════════════════════════════════════════════════════
# RAW TABLE EXPORT
# ══════════════════════════════════════════════════════════════════
def export_raw_table(spark, wb, view_name, sheet_name, label):
    print(f"\n   📋 Exporting {label} → '{sheet_name}' ...")
    try:
        sdf = spark.sql(f"SELECT * FROM {view_name}")
        pdf = safe_to_pandas(sdf)

        ws = wb.create_sheet(title=sheet_name)
        write_dataframe(ws, pdf, start_row=1)
        print(f"      → {len(pdf):,} rows × {len(pdf.columns)} columns")
        return {"available": True,
                "row_count": len(pdf),
                "col_count": len(pdf.columns)}

    except Exception as e:
        ws = wb.create_sheet(title=sheet_name)
        ws["A1"] = f"⚠️  Table '{view_name}' could not be exported: {e}"
        ws["A1"].font = Font(name="Arial", size=10,
                             italic=True, color="C0392B")
        print(f"      ❌ Failed: {e}")
        return {"available": False, "row_count": 0, "col_count": 0}


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    print("\n" + "=" * 60)
    print("  🏪 E-COMMERCE ANALYTICS PLATFORM")
    print("  Raw Data Export → Excel")
    print("=" * 60)

    print("\n📡 Starting Spark...")
    spark = create_spark()
    print(f"   Spark version: {spark.version}")

    row_counts  = load_tables(spark)
    export_time = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    wb          = openpyxl.Workbook()

    EXPORTS = [
        ("orders",           "Orders_Raw",          "Orders"),
        ("payments",         "Payments_Raw",         "Payments"),
        ("user_events",      "User_Events_Raw",      "User Events"),
        ("customer_rfm",     "Customer_RFM_Raw",     "Customer RFM"),
        ("churn_candidates", "Churn_Candidates_Raw", "Churn Candidates"),
    ]

    print("\n" + "=" * 60)
    print("  Exporting raw tables to Excel")
    print("=" * 60)

    table_meta = {}
    for view, sheet, label in EXPORTS:
        if row_counts.get(view, 0) == 0:
            ws = wb.create_sheet(title=sheet)
            ws["A1"] = f"⚠️  Table '{view}' not available or empty."
            ws["A1"].font = Font(name="Arial", size=10,
                                 italic=True, color="C0392B")
            table_meta[view] = {"available": False,
                                 "row_count": 0, "col_count": 0}
            print(f"\n   ⚠️  {label} — skipped (empty / not found)")
        else:
            table_meta[view] = export_raw_table(
                spark, wb, view, sheet, label)

    # Index sheet — write into default first sheet, move to front
    write_index_sheet(wb, table_meta, export_time)
    wb.move_sheet("Index", offset=-len(wb.sheetnames) + 1)

    wb.save(EXCEL_PATH)

    # ── Summary ────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  ✅ EXPORT COMPLETE")
    print("=" * 60)
    total_rows = sum(
        m["row_count"] for m in table_meta.values() if m["available"]
    )
    print(f"\n📗 File  : {EXCEL_PATH}")
    print(f"   Rows  : {total_rows:,} total across all sheets")
    print(f"   Sheets: {len(wb.sheetnames)}\n")
    for view, sheet, _ in EXPORTS:
        m = table_meta[view]
        status = (f"{m['row_count']:,} rows × {m['col_count']} cols"
                  if m["available"] else "not available")
        icon = "✅" if m["available"] else "⚠️ "
        print(f"   {icon}  {sheet:<26} {status}")

    spark.stop()
    print("\n🎉 Done!")


if __name__ == "__main__":
    main()
